#!/usr/bin/env python3
"""
Outil de consultation et export des factures extraites depuis DynamoDB.

Usage:
    python view_invoices.py                         # Affiche toutes les factures
    python view_invoices.py --export                # Export XLSX (invoices_export.xlsx)
    python view_invoices.py --export --out mon.xlsx # Export dans un fichier personnalisé
    python view_invoices.py --fournisseur ORANGE    # Filtrer par fournisseur
    python view_invoices.py --depuis 2024-01-01     # Filtrer depuis une date
"""

import argparse
import boto3
import io
import json
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

TABLE_NAME = os.getenv("DYNAMODB_TABLE_NAME", "invoices-extractor")
REGION     = os.getenv("AWS_REGION", "us-west-2")

COLUMNS = [
    ("nom_fichier",      "Fichier"),
    ("fournisseur",      "Fournisseur"),
    ("numero_facture",   "N° Facture"),
    ("date_facture",     "Date"),
    ("montant_ht",       "Montant HT"),
    ("devise",           "Devise"),
    ("chrono",           "Chrono"),
    ("couverture",       "Couverture"),
    ("extraction_date",  "Extrait le"),
]


def scan_table(fournisseur=None, depuis=None):
    dynamodb = boto3.client("dynamodb", region_name=REGION)
    paginator = dynamodb.get_paginator("scan")

    items = []
    for page in paginator.paginate(TableName=TABLE_NAME):
        items.extend(page.get("Items", []))

    records = []
    for item in items:
        raw = item.get("raw_data", {}).get("S")
        if raw:
            try:
                data = json.loads(raw)
            except json.JSONDecodeError:
                data = {}
        else:
            data = {}

        # Fusionner les champs directs de DynamoDB avec raw_data
        for key, val in item.items():
            if key == "raw_data":
                continue
            dtype = list(val.keys())[0]
            data.setdefault(key, val[dtype])

        records.append(data)

    # Filtres
    if fournisseur:
        f = fournisseur.upper()
        records = [r for r in records if f in str(r.get("fournisseur", "")).upper()]

    if depuis:
        records = [r for r in records if str(r.get("date_facture", "")) >= depuis]

    # Tri par date décroissante
    records.sort(key=lambda r: str(r.get("date_facture") or ""), reverse=True)
    return records


def display_table(records):
    if not records:
        print("Aucune facture trouvée.")
        return

    # Calcul des largeurs de colonnes
    widths = {key: len(label) for key, label in COLUMNS}
    for r in records:
        for key, _ in COLUMNS:
            widths[key] = max(widths[key], len(str(r.get(key) or "")))

    sep = "+-" + "-+-".join("-" * widths[k] for k, _ in COLUMNS) + "-+"
    header = "| " + " | ".join(
        label.ljust(widths[key]) for key, label in COLUMNS
    ) + " |"

    print(sep)
    print(header)
    print(sep)
    for r in records:
        row = "| " + " | ".join(
            str(r.get(key) or "").ljust(widths[key]) for key, _ in COLUMNS
        ) + " |"
        print(row)
    print(sep)
    print(f"\n{len(records)} facture(s) trouvée(s).")

    # Total montant HT
    total = 0
    for r in records:
        try:
            total += float(r.get("montant_ht") or 0)
        except (ValueError, TypeError):
            pass
    if total:
        print(f"Total HT : {total:,.2f}")


def export_xlsx(records, output_path):
    if not records:
        print("Aucune donnée à exporter.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Factures"

    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(bold=True, color="FFFFFF")

    fields  = [key   for key, _     in COLUMNS]
    headers = [label for _,   label in COLUMNS]

    # En-tête
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Données
    for row_idx, record in enumerate(records, start=2):
        for col_idx, field in enumerate(fields, start=1):
            raw  = record.get(field) or ""
            cell = ws.cell(row=row_idx, column=col_idx)

            if field == "montant_ht" and raw:
                try:
                    cell.value         = float(raw)
                    cell.number_format = '#,##0.00'
                    cell.alignment     = Alignment(horizontal="right")
                    continue
                except (ValueError, TypeError):
                    pass

            if field == "date_facture" and raw:
                try:
                    cell.value         = datetime.strptime(raw, "%Y-%m-%d")
                    cell.number_format = 'DD/MM/YYYY'
                    cell.alignment     = Alignment(horizontal="center")
                    continue
                except ValueError:
                    pass

            if field == "extraction_date" and raw:
                try:
                    cell.value         = datetime.fromisoformat(raw).replace(tzinfo=None)
                    cell.number_format = 'DD/MM/YYYY HH:MM'
                    cell.alignment     = Alignment(horizontal="center")
                    continue
                except ValueError:
                    pass

            cell.value = raw

    # Largeurs automatiques
    for col_idx, field in enumerate(fields, start=1):
        max_len = len(headers[col_idx - 1])
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = "A2"

    wb.save(output_path)
    print(f"Export XLSX : {output_path} ({len(records)} ligne(s))")


def main():
    parser = argparse.ArgumentParser(description="Consulter et exporter les factures extraites")
    parser.add_argument("--export", action="store_true", help="Exporter en XLSX")
    parser.add_argument("--out", default="invoices_export.xlsx", help="Fichier XLSX de sortie")
    parser.add_argument("--fournisseur", help="Filtrer par fournisseur (ex: ORANGE)")
    parser.add_argument("--depuis", help="Filtrer depuis une date YYYY-MM-DD (ex: 2024-01-01)")
    args = parser.parse_args()

    print(f"Connexion à DynamoDB ({TABLE_NAME}, {REGION})...")
    try:
        records = scan_table(fournisseur=args.fournisseur, depuis=args.depuis)
    except Exception as e:
        print(f"Erreur : {e}")
        sys.exit(1)

    display_table(records)

    if args.export:
        export_xlsx(records, args.out)


if __name__ == "__main__":
    main()
