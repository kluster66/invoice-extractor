#!/usr/bin/env python3
"""
Interface graphique NiceGui pour consulter et exporter les factures extraites.
Lancement : python ui_invoices.py
"""

import boto3
import io
import json
import os
from datetime import datetime
from nicegui import ui, app, events
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

TABLE_NAME  = os.getenv("DYNAMODB_TABLE_NAME", "invoices-extractor")
S3_BUCKET   = os.getenv("S3_INPUT_BUCKET", "")
REGION      = os.getenv("AWS_REGION", "us-west-2")

COLUMNS = [
    {"name": "nom_fichier",     "label": "Fichier",      "field": "nom_fichier",     "sortable": True},
    {"name": "fournisseur",     "label": "Fournisseur",  "field": "fournisseur",     "sortable": True},
    {"name": "numero_facture",  "label": "N° Facture",   "field": "numero_facture",  "sortable": True},
    {"name": "date_facture",    "label": "Date",         "field": "date_facture",    "sortable": True},
    {"name": "montant_ht",      "label": "Montant HT",   "field": "montant_ht",      "sortable": True},
    {"name": "devise",          "label": "Devise",       "field": "devise",          "sortable": True},
    {"name": "chrono",          "label": "Chrono",       "field": "chrono",          "sortable": True},
    {"name": "couverture",      "label": "Couverture",   "field": "couverture",      "sortable": True},
    {"name": "extraction_date", "label": "Extrait le",   "field": "extraction_date", "sortable": True},
]


# ---------------------------------------------------------------------------
# DynamoDB helpers
# ---------------------------------------------------------------------------

def fetch_records() -> list[dict]:
    dynamodb = boto3.client("dynamodb", region_name=REGION)
    paginator = dynamodb.get_paginator("scan")
    items = []
    for page in paginator.paginate(TableName=TABLE_NAME):
        items.extend(page.get("Items", []))

    records = []
    for item in items:
        raw = item.get("raw_data", {}).get("S")
        data = {}
        if raw:
            try:
                data = json.loads(raw)
            except json.JSONDecodeError:
                pass
        for key, val in item.items():
            if key == "raw_data":
                continue
            dtype = list(val.keys())[0]
            data.setdefault(key, val[dtype])

        row = {col["field"]: str(data.get(col["field"]) or "") for col in COLUMNS}
        # Conserver invoice_id pour la suppression
        row["invoice_id"] = item.get("invoice_id", {}).get("S", "")
        records.append(row)

    records.sort(key=lambda r: r.get("date_facture") or "", reverse=True)
    return records


def delete_records(invoice_ids: list[str]) -> int:
    dynamodb = boto3.client("dynamodb", region_name=REGION)
    count = 0
    for iid in invoice_ids:
        dynamodb.delete_item(
            TableName=TABLE_NAME,
            Key={"invoice_id": {"S": iid}},
        )
        count += 1
    return count


# ---------------------------------------------------------------------------
# S3 helpers
# ---------------------------------------------------------------------------

def upload_to_s3(filename: str, content: bytes) -> None:
    s3 = boto3.client("s3", region_name=REGION)
    s3.put_object(Bucket=S3_BUCKET, Key=filename, Body=content)


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------

HEADER_FILL  = PatternFill("solid", fgColor="1E40AF")   # bleu
HEADER_FONT  = Font(bold=True, color="FFFFFF")
NUMBER_FMT   = '#,##0.00'
DATE_FMT     = 'DD/MM/YYYY'
DATETIME_FMT = 'DD/MM/YYYY HH:MM'

def build_xlsx(records: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Factures"

    headers = [c["label"] for c in COLUMNS]
    fields  = [c["field"] for c in COLUMNS]

    # En-tête
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Données
    for row_idx, record in enumerate(records, start=2):
        for col_idx, field in enumerate(fields, start=1):
            raw = record.get(field) or ""
            cell = ws.cell(row=row_idx, column=col_idx)

            if field == "montant_ht" and raw:
                try:
                    cell.value          = float(raw)
                    cell.number_format  = NUMBER_FMT
                    cell.alignment      = Alignment(horizontal="right")
                    continue
                except (ValueError, TypeError):
                    pass

            if field == "date_facture" and raw:
                try:
                    cell.value         = datetime.strptime(raw, "%Y-%m-%d")
                    cell.number_format = DATE_FMT
                    cell.alignment     = Alignment(horizontal="center")
                    continue
                except ValueError:
                    pass

            if field == "extraction_date" and raw:
                try:
                    cell.value         = datetime.fromisoformat(raw).replace(tzinfo=None)
                    cell.number_format = DATETIME_FMT
                    cell.alignment     = Alignment(horizontal="center")
                    continue
                except ValueError:
                    pass

            cell.value = raw

    # Largeurs automatiques
    for col_idx, field in enumerate(fields, start=1):
        max_len = len(headers[col_idx - 1])
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    # Filtre auto + ligne figée
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()



# ---------------------------------------------------------------------------
# Page principale
# ---------------------------------------------------------------------------

@ui.page("/")
def main_page():
    all_records: list[dict] = []
    filtered:    list[dict] = []

    # ── En-tête ──────────────────────────────────────────────────────────────
    with ui.header().classes("items-center justify-between bg-blue-800 text-white px-6 py-3"):
        ui.label("Invoice Extractor — Factures").classes("text-xl font-bold")
        ui.button("Quitter", icon="power_settings_new", on_click=lambda: app.shutdown()).props("flat color=white")

    with ui.column().classes("w-full p-6 gap-6"):

        # ── Section upload ────────────────────────────────────────────────────
        with ui.card().classes("w-full"):
            ui.label("Déposer des factures").classes("text-base font-semibold mb-2")

            upload_status = ui.label("").classes("text-sm text-gray-500")

            async def handle_upload(e: events.UploadEventArguments):
                try:
                    filename = e.file.name
                    content  = await e.file.read()
                    upload_status.set_text(f"Upload en cours : {filename}…")
                    upload_to_s3(filename, content)
                    ui.notify(f"✅ {filename} uploadé — la Lambda va traiter la facture.", type="positive", timeout=5000)
                    upload_status.set_text(f"Dernier upload réussi : {filename}")
                except Exception as exc:
                    ui.notify(f"❌ Erreur upload : {exc}", type="negative", timeout=0)
                    upload_status.set_text(f"Erreur : {exc}")

            def handle_rejected(e):
                ui.notify("❌ Fichier refusé (format non supporté).", type="warning")

            ui.upload(
                label="Cliquez ou glissez vos PDF ici",
                multiple=True,
                auto_upload=True,
                on_upload=handle_upload,
                on_rejected=handle_rejected,
            ).props('accept=".pdf,.PDF,application/pdf" color=blue flat bordered').classes("w-full")

        # ── Filtres ───────────────────────────────────────────────────────────
        with ui.card().classes("w-full"):
            ui.label("Filtres").classes("text-base font-semibold mb-2")
            with ui.row().classes("items-end gap-4 flex-wrap"):
                fournisseur_input = ui.input("Fournisseur").props("clearable outlined dense").classes("w-48")
                depuis_input      = ui.input("Depuis (YYYY-MM-DD)").props("clearable outlined dense").classes("w-48")
                ui.button("Filtrer",       icon="search", on_click=lambda: apply_filters()).props("color=blue")
                ui.button("Réinitialiser", icon="clear",  on_click=lambda: reset_filters()).props("flat")

        # ── Résumé + actions ──────────────────────────────────────────────────
        with ui.row().classes("items-center justify-between w-full"):
            summary = ui.label("").classes("text-sm text-gray-500")
            with ui.row().classes("gap-2"):
                ui.button("Exporter la sélection", icon="download",
                          on_click=lambda: export_selected()).props("color=blue flat")
                ui.button("Tout exporter", icon="download",
                          on_click=lambda: export_all()).props("color=blue outline")
                ui.button("Supprimer la sélection", icon="delete",
                          on_click=lambda: confirm_delete_selected()).props("color=red flat")
                ui.button("Tout supprimer", icon="delete_forever",
                          on_click=lambda: confirm_delete_all()).props("color=red outline")

        # ── Tableau ───────────────────────────────────────────────────────────
        # Colonnes affichées (sans invoice_id)
        table_cols = [{"name": "select", "label": "", "field": "select"}] + COLUMNS
        table = ui.table(columns=table_cols, rows=[], row_key="invoice_id",
                         selection="multiple").classes("w-full")
        table.props("flat bordered dense virtual-scroll")

    # ── Logique export ────────────────────────────────────────────────────────

    def export_selected():
        selected = table.selected
        if not selected:
            ui.notify("Aucune ligne sélectionnée.", type="warning")
            return
        ui.download(build_xlsx(selected), "selection_export.xlsx")

    def export_all():
        if not filtered:
            ui.notify("Aucune donnée à exporter.", type="warning")
            return
        ui.download(build_xlsx(filtered), "invoices_export.xlsx")

    # ── Logique filtres ───────────────────────────────────────────────────────

    def apply_filters():
        f = fournisseur_input.value.strip().upper() if fournisseur_input.value else ""
        d = depuis_input.value.strip()              if depuis_input.value      else ""
        result = all_records
        if f:
            result = [r for r in result if f in r.get("fournisseur", "").upper()]
        if d:
            result = [r for r in result if r.get("date_facture", "") >= d]
        filtered.clear()
        filtered.extend(result)
        table.rows = filtered
        table.update()
        refresh_summary()

    def reset_filters():
        fournisseur_input.set_value("")
        depuis_input.set_value("")
        filtered.clear()
        filtered.extend(all_records)
        table.rows = all_records
        table.update()
        refresh_summary()

    def refresh_summary():
        total = 0
        for r in filtered:
            try:
                total += float(r.get("montant_ht") or 0)
            except (ValueError, TypeError):
                pass
        total_str = f"  —  Total HT : {total:,.2f}" if total else ""
        summary.set_text(f"{len(filtered)} facture(s){total_str}")

    # ── Logique suppression ───────────────────────────────────────────────────

    def confirm_delete_selected():
        selected = table.selected
        if not selected:
            ui.notify("Aucune ligne sélectionnée.", type="warning")
            return
        with ui.dialog() as dlg, ui.card():
            ui.label(f"Supprimer {len(selected)} facture(s) sélectionnée(s) ?").classes("font-semibold")
            ui.label("Cette action est irréversible.").classes("text-sm text-gray-500")
            with ui.row().classes("justify-end gap-2 mt-4"):
                ui.button("Annuler",    on_click=dlg.close).props("flat")
                ui.button("Supprimer",  on_click=lambda: do_delete_selected(selected, dlg)).props("color=red")
        dlg.open()

    def do_delete_selected(selected, dlg):
        dlg.close()
        ids = [r["invoice_id"] for r in selected if r.get("invoice_id")]
        try:
            n = delete_records(ids)
            ui.notify(f"✅ {n} facture(s) supprimée(s).", type="positive")
        except Exception as exc:
            ui.notify(f"❌ Erreur suppression : {exc}", type="negative")
        load_data()

    def confirm_delete_all():
        if not all_records:
            ui.notify("Aucune donnée à supprimer.", type="warning")
            return
        with ui.dialog() as dlg, ui.card():
            ui.label(f"Supprimer TOUTES les {len(all_records)} factures ?").classes("font-semibold text-red-600")
            ui.label("Cette action est irréversible.").classes("text-sm text-gray-500")
            with ui.row().classes("justify-end gap-2 mt-4"):
                ui.button("Annuler",         on_click=dlg.close).props("flat")
                ui.button("Tout supprimer",  on_click=lambda: do_delete_all(dlg)).props("color=red")
        dlg.open()

    def do_delete_all(dlg):
        dlg.close()
        ids = [r["invoice_id"] for r in all_records if r.get("invoice_id")]
        try:
            n = delete_records(ids)
            ui.notify(f"✅ {n} facture(s) supprimée(s).", type="positive")
        except Exception as exc:
            ui.notify(f"❌ Erreur suppression : {exc}", type="negative")
        load_data()

    # ── Chargement initial ────────────────────────────────────────────────────

    def load_data():
        summary.set_text("Chargement...")
        try:
            records = fetch_records()
            all_records.clear()
            all_records.extend(records)
            filtered.clear()
            filtered.extend(records)
            table.rows = records
            table.selected = []
            table.update()
            refresh_summary()
        except Exception as e:
            ui.notify(f"Erreur DynamoDB : {e}", type="negative")
            summary.set_text("Erreur de chargement")

    ui.timer(0.1, load_data, once=True)


app.on_disconnect(lambda: app.shutdown())

ui.run(title="Invoice Extractor", port=8080, reload=False)
